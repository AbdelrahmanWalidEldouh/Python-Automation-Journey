import openpyxl

wb = openpyxl.load_workbook("example.xlsx")
sheet = wb["Sheet1"]
print(sheet.max_row)
print(sheet.max_column)
