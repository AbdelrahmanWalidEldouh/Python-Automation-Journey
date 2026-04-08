from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = load_workbook("store_data.xlsx")
sheet = wb["Data"]
# تغيير الخط (B1)
sheet["B1"].font = Font(bold=True, color="FF0000", size=14)

# تلوين خلفية الخلية
sheet["A1"].fill = PatternFill(
    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
)

# محاذاة النص للوسط
sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

# دمج الخلايا
sheet.merge_cells("A5:B5")
sheet["A5"] = "Total Summary"

wb.save("store_data.xlsx")
